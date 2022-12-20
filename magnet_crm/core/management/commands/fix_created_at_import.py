
from django.core.management.base import BaseCommand, CommandError

from django.contrib.auth.models import User
from django.db import IntegrityError, transaction
from datetime import datetime
import requests

from django.core import management
from django.core.management import call_command
from django.utils import timezone


from core.models import *
import mysql.connector
from mysql.connector import errorcode
from client.models import *
from staff.models import * 

import openpyxl
import os


class Command(BaseCommand):
	help = 'check_user_journey_deposit'

	# def add_arguments(self, parser):
	#     parser.add_argument('poll_ids', nargs='+', type=int)

	# def add_arguments(self, parser):
		# parser.add_argument('username', type=str, help='username & password')

	def handle(self, *args, **kwargs):
		directory = os.getcwd()
		directory += '/core/management/commands/'
		print(directory, type(directory))
		filename = 'rahma_febrina@magnetfx.co.id.xlsx'
		directory += filename
		print(directory)
		wb = openpyxl.load_workbook(filename = directory)
		print(wb)
		try:
			with transaction.atomic():
				staff = Staff.objects.filter(profile__user__email='rahma_febrina@magnetfx.co.id').first()
				external_data_wb = wb["Data Eksternal"]
				for row in external_data_wb.iter_rows(2, external_data_wb.max_row):
					if row[0].value != None:
						# print("Tanggal data diberikan: ", row[1].value)
						# print("Nama: ", row[2].value)
						# print("No Telepon: ", row[3].value)
						# print("Email: ", row[4].value)
						share_date = row[1].value
						nama = row[2].value
						nomor_telepon = row[3].value
						email = row[4].value
						phone_no = str(row[3].value)
						if phone_no != None and phone_no != '':
							phone_no = str(row[3].value).strip()
							phone_no = phone_no.replace('-', '')

							phone_no = phone_no.encode('ascii', 'ignore').decode("utf-8")

							if len(phone_no) > 0:
								if phone_no[0] == '8':
									phone_no = '62' + phone_no


						print(nama, nomor_telepon, email)
						existing_client = Client.objects.get(nama=nama, phone_no=phone_no, email=email)
						existing_client.created_at = share_date
						existing_client.save()
						print(existing_client)

				leads_data_wb = wb["Data Leads"]
				for row in leads_data_wb.iter_rows(2, leads_data_wb.max_row):
					if row[0].value != None:
						# print("Tanggal data diberikan: ", row[1].value)
						# print("Nama: ", row[2].value)
						# print("No Telepon: ", row[3].value)
						# print("Email: ", row[4].value)
						share_date = row[1].value
						nama = row[2].value
						nomor_telepon = row[3].value
						email = row[4].value
						phone_no = str(row[3].value)
						if phone_no != None and phone_no != '':
							phone_no = str(row[3].value).strip()
							phone_no = phone_no.replace('-', '')

							phone_no = phone_no.encode('ascii', 'ignore').decode("utf-8")

							if len(phone_no) > 0:
								if phone_no[0] == '8':
									phone_no = '62' + phone_no


						print(nama, nomor_telepon, email)
						existing_client = Client.objects.get(nama=nama, phone_no=phone_no, email=email)
						existing_client.created_at = share_date
						existing_client.save()
						print(existing_client)
				
						
				print('start', external_data_wb)
				print(staff)

		except:
			print('error bung')