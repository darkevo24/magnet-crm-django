from django.shortcuts import render
# # from core.forms.admin.main.main import *
# # from referal.models import *
from django.contrib.auth.decorators import login_required

from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import JsonResponse, HttpResponseRedirect
from client.forms import *
from django.db import IntegrityError, transaction
from django.contrib import messages
from django.core.paginator import Paginator

from datetime import datetime
from django.utils import timezone
import re

from core.forms.main import *
from client.models import *
from staff.models import *
from client.forms import *
from client.utils import *
from xlrd import open_workbook ,xldate_as_tuple
from magnet_crm.task import *
from datetime import datetime, timedelta
from django.utils.timezone import make_aware
import csv
import openpyxl
from io import StringIO
from decimal import *
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from core.datatable.server_side_datatable_view import ServerSideDatatableView
from core.datatable.client_suspect_datatable_view import ClientSuspectServerSideDatatableView
from django.views.decorators.csrf import csrf_exempt

CLEANR = re.compile('<.*?>') 

@login_required
def client_import(request):
	template = 'admin/client/client_import.html'
	import_form = ClientImportForm(request.POST or None, request.FILES or None)
	jakarta_timezone = pytz.timezone('Asia/Jakarta')
	print(request.POST)
	
	if request.POST:
		try:
			with transaction.atomic():

				if import_form.is_valid():
					excel_file = request.FILES["file"]
					print(excel_file)
					staff_email = excel_file.name.split('.xlsx')[0]
					selected_staff = Staff.objects.filter(id=import_form.cleaned_data['staff']).first()
					print('selected_staff', selected_staff)
					wb = openpyxl.load_workbook(excel_file)
					external_data_wb = wb["Data Eksternal"]
					leads_data_wb = wb["Data Leads"]

					# Langsung dibuat client
					for row in external_data_wb.iter_rows(2, external_data_wb.max_row):

						if row[0].value != None:
							# print("Tanggal data diberikan: ", row[1].value)
							# print("Nama: ", row[2].value)
							# print("No Telepon: ", row[3].value)
							# print("Email: ", row[4].value)
							# print("Source: ", row[5].value)
							# print("Medium: ", row[6].value)
							# print("Campaign: ", row[7].value)
							# print("GCLID: ", row[8].value)
							# print("Category data: ", row[9].value)
							# print("Tanggal respon: ", row[10].value)
							phone_no = str(row[3].value)
							if phone_no != None and phone_no != '':
								phone_no = str(row[3].value).strip()
								phone_no = phone_no.replace('-', '')

								phone_no = phone_no.encode('ascii', 'ignore').decode("utf-8")

								if len(phone_no) > 0:
									if phone_no[0] == '8':
										phone_no = '62' + phone_no
							print('****', row[2].value, phone_no)
							existing_client = Client.objects.filter(email=row[4].value, nama=row[2].value, phone_no=phone_no).first()
							if existing_client == None:
								new_client = Client()
								new_client.created_at = row[1].value.replace(tzinfo=jakarta_timezone)
								new_client.created_by = request.user
								new_client.nama = row[2].value
								print('+++++', row[2].value, phone_no)
								new_client.phone_no = phone_no
								new_client.email = row[4].value
								new_client.source = '2'
								if 5 in row:
									if row[5].value == "google":
										new_client.source_detail_1 = None
										new_client.source_detail_2 = "4"
									elif row[5].value == "fb / ig":
										new_client.source_detail_1 = None
										new_client.source_detail_2 = "1"
									else:
										new_client.source_detail_1 = None
								if 6 in row:
									new_client.medium = row[6].value
								if 7 in row:
									new_client.campaign = row[7].value
								if 8 in row:
									new_client.gclid = row[8].value
								new_client.save()

								if selected_staff is not None and selected_staff is not "":
									c_staff = Client_Staff()
									c_staff.client = new_client
									c_staff.staff =	selected_staff
									c_staff.created_by = request.user				
									c_staff.save()
							else:
								client_staff = Client_Staff.objects.filter(is_active=True, client=existing_client, staff=selected_staff).first()
								if client_staff != None :
									print('client_staff tidak sama dengan None')
									if client_staff.staff.profile.email != selected_staff.profile.user.email:
										print('client_staff email tidak sama', client_staff.staff.profile.email, selected_staff.profile.user.email)
										client_staff.is_active = False
										client_staff.save()

										new_client_staff = Client_Staff()
										new_client_staff.client = existing_client
										new_client_staff.staff = selected_staff
										new_client_staff.updated_by = new_client_staff.created_by = request.user
										print('new_client_staff saved !')
										new_client_staff.save()

										client_journey_list = Client_Journey.objects.filter(client=existing_client, is_active=True)
										for client_journey in client_journey_list:
											client_journey.is_active = False
											client_journey.updated_by = request.user
											client_journey.save()

											new_client_journey = Client_Journey()
											new_client_journey.client = existing_client
											new_client_journey.staff = selected_staff
											new_client_journey.journal_type = client_journey.journal_type
											new_client_journey.created_by = request.user
											new_client_journey.updated_by = request.user
											new_client_journey.created_at = client_journey.created_at
											new_client_journey.save()
								else:
									print('client_staff sama dengan None')
									new_client_staff = Client_Staff()
									new_client_staff.client = existing_client
									new_client_staff.staff = selected_staff
									new_client_staff.updated_by = new_client_staff.created_by = request.user
									print('new_client_staff saved !!')
									new_client_staff.save()

					# Cek client udah exist belom, query email
					for row in leads_data_wb.iter_rows(2, leads_data_wb.max_row):
						if row[0].value != None:
							# print("Tanggal data: ", row[1].value)
							# print("Nama: ", row[2].value)
							# print("No Telepon: ", row[3].value)
							# print("Email: ", row[4].value)
							# print("Source: ", row[8].value)
							# print("Medium: ", row[9].value)
							# print("Campaign: ", row[10].value)
							# print("AE code: ", row[11].value)
							# print("User code: ", row[12].value)
							# print("Category data: ", row[13].value)
							# print("Tanggal respon: ", row[14].value)
							phone_no = str(row[3].value)
							if phone_no != None and phone_no != '':
								phone_no = str(row[3].value).strip()
								phone_no = phone_no.replace('-', '')
								phone_no = phone_no.encode('ascii', 'ignore').decode("utf-8")
								if len(phone_no) > 0:
									if phone_no[0] == '8':
										phone_no = '62' + phone_no
							
							existing_client = Client.objects.filter(email=row[4].value, nama=row[2].value, phone_no=phone_no, is_active=True).first()
							print("existing leads data: ", existing_client)
							if existing_client != None:
								print("existing client found")
								
								# existing_client.is_active = False
								client_staff = Client_Staff.objects.filter(is_active=True, client=existing_client, staff=selected_staff).first()
								if client_staff != None :
									print('client_staff tidak sama dengan None')
									if client_staff.staff.profile.email != selected_staff.profile.user.email:
										print('client_staff email tidak sama', client_staff.staff.profile.email, selected_staff.profile.user.email)
										client_staff.is_active = False
										client_staff.save()

										new_client_staff = Client_Staff()
										new_client_staff.client = existing_client
										new_client_staff.staff = selected_staff
										new_client_staff.updated_by = new_client_staff.created_by = request.user
										print('new_client_staff saved !')
										new_client_staff.save()

										client_journey_list = Client_Journey.objects.filter(client=existing_client, is_active=True)
										for client_journey in client_journey_list:
											client_journey.is_active = False
											client_journey.updated_by = request.user
											client_journey.save()

											new_client_journey = Client_Journey()
											new_client_journey.client = existing_client
											new_client_journey.staff = selected_staff
											new_client_journey.journal_type = client_journey.journal_type
											new_client_journey.created_by = request.user
											new_client_journey.updated_by = request.user
											new_client_journey.created_at = client_journey.created_at
											new_client_journey.save()
								else:
									print('client_staff sama dengan None')
									new_client_staff = Client_Staff()
									new_client_staff.client = existing_client
									new_client_staff.staff = selected_staff
									new_client_staff.updated_by = new_client_staff.created_by = request.user
									print('new_client_staff saved !!')
									new_client_staff.save()

							else:
								print("existing client not found")
								new_client = Client()
								new_client.created_at = row[1].value.replace(tzinfo=jakarta_timezone)
								new_client.created_by = request.user
								new_client.nama = row[2].value
								
								print('____', row[2].value, phone_no)
								new_client.phone_no = phone_no
								new_client.email = row[4].value
								new_client.source = '0'
								if row[5].value == "google":
									new_client.source_detail_1 = None
									new_client.source_detail_2 = "4"
								elif row[5].value == "fb / ig":
									new_client.source_detail_1 = None
									new_client.source_detail_2 = "1"
								else:
									new_client.source_detail_1 = None
								new_client.medium = row[6].value
								new_client.campaign = row[7].value
								client_aecode = row[8].value
								if client_aecode == None:
									client_aecode = ''
								new_client.aecode = client_aecode
								print('sebelum save')	
								new_client.save()		
								print('setelah save')				

								new_client_staff = Client_Staff()
								new_client_staff.client = new_client
								new_client_staff.staff = selected_staff
								new_client_staff.updated_by = new_client_staff.created_by = request.user
								print('new_client_staff saved !!!!')
								new_client_staff.save()
								

					# counter = 0 
					# for row in file:
					# 	# print(row)
					# 	if counter == 0:
					# 		counter+=1
					# 		pass
					# 	else:
					# 		print("temp",row)

					# 		temp = row.split(";")
					# 		print("temp",temp)
					# 		new_client = Client()
					# 		new_client.nama = temp[2]
					# 		try:
					# 			new_client.phone_no = temp[3]
					# 		except:
					# 			new_client.phone_no = 0
					# 		new_client.email = temp[4]

					# 		if temp[5] == "google":

					# 			new_client.source_detail_1 = None
					# 			new_client.source_detail_2 = "4"
					# 		elif temp[5] == "fb / ig":
					# 			new_client.source_detail_1 = None
					# 			new_client.source_detail_2 = "1"
					# 		else:
					# 			new_client.source_detail_1 = None

					# 		new_client.created_by = request.user
					# 		new_client.save()

					# 		selected_staff = Staff.objects.filter(id=import_form.cleaned_data['staff']).first()
					# 		if selected_staff is not None and selected_staff is not "":
								
					# 			c_staff = Client_Staff()
					# 			c_staff.client = new_client
					# 			c_staff.staff =	selected_staff
					# 			c_staff.created_by = request.user				
					# 			c_staff.save()




					
					# 			ini kalo pake xls / xlsx
					# 			excel_file = import_form.cleaned_data['file']
					# 			rb = open_workbook(file_contents=excel_file.read())
					# 			sheet = rb.sheet_by_index(0)
					# 			print(sheet.nrows)
					# 			for x in range(sheet.nrows):
					# 				print(sheet.row_values(x)[0])

					messages.success(request, 'Berhasil Impor Data.')
					return redirect(reverse('client-import'))
				else:
					print(import_form.errors)
					print(import_form.errors)

		except IntegrityError as e:
			print(e)

	context = {
		'import_form': import_form,
		'menu':'client_import',
	}
	return render(request,template,context=context)

@login_required
def client_share_date(request):
	template = 'admin/client/client_import.html'
	import_form = ClientImportForm(request.POST or None, request.FILES or None)
	jakarta_timezone = pytz.timezone('Asia/Jakarta')
	print(request.POST)
	
	if request.POST:
		try:
			with transaction.atomic():

				if import_form.is_valid():
					excel_file = request.FILES["file"]
					print(excel_file)
					staff_email = excel_file.name.split('.xlsx')[0]
					
					wb = openpyxl.load_workbook(excel_file)
					external_data_wb = wb["Data Eksternal"]
					leads_data_wb = wb["Data Leads"]
					for row in external_data_wb.iter_rows(2, external_data_wb.max_row):
						if row[0].value != None:
							share_date = row[1].value
							nama = row[2].value
							nomor_telepon = row[3].value
							email = row[4].value
							phone_no = str(row[3].value)
							if phone_no != None and phone_no != '':
								phone_no = str(row[3].value).strip()
								phone_no = phone_no.replace('-', '')

								phone_no = phone_no.encode('ascii', 'ignore').decode("utf-8")

								if len(phone_no) > 0:
									if phone_no[0] == '8':
										phone_no = '62' + phone_no


							print(nama, nomor_telepon, email)
							existing_client = Client.objects.filter(nama=nama, phone_no=phone_no, email=email).order_by('-id')
							if existing_client.count() > 0:
								existing_client = existing_client[0]
								existing_client.created_at = share_date.replace(tzinfo=jakarta_timezone)
								existing_client.save()
								print(existing_client)
					
					for row in leads_data_wb.iter_rows(2, leads_data_wb.max_row):
						if row[0].value != None:
							# print("Tanggal data diberikan: ", row[1].value)
							# print("Nama: ", row[2].value)
							# print("No Telepon: ", row[3].value)
							# print("Email: ", row[4].value)
							share_date = row[1].value
							nama = row[2].value
							nomor_telepon = row[3].value
							email = row[4].value
							phone_no = str(row[3].value)
							if phone_no != None and phone_no != '':
								phone_no = str(row[3].value).strip()
								phone_no = phone_no.replace('-', '')

								phone_no = phone_no.encode('ascii', 'ignore').decode("utf-8")

								if len(phone_no) > 0:
									if phone_no[0] == '8':
										phone_no = '62' + phone_no

							existing_client = Client.objects.filter(nama=nama, phone_no=phone_no, email=email).order_by('-id')
							if existing_client.count() > 0:
								existing_client = existing_client[0]
								existing_client.created_at = share_date.replace(tzinfo=jakarta_timezone)
								existing_client.save()
								print(existing_client)
					
					messages.success(request, 'Berhasil Impor Data.')
					return redirect(reverse('client-share-date'))
				else:
					print(import_form.errors)
					print(import_form.errors)

		except IntegrityError as e:
			print(e)

	context = {
		'import_form': import_form,
		'menu':'client_import',
	}
	return render(request,template,context=context)


@login_required
def client_own_suspect_list(request):
	now = timezone.now()
	month = request.GET.get('month')
	year = request.GET.get('year') or now.year

	staff = Staff.objects.filter(profile__user=request.user).first()


	client_list = Client_Staff.objects.filter(is_active=True,is_own_client_suspect=True)
	template = 'admin/client/suspect/client_own_list.html'
	if month != "" and month != None:
		client_list = Client_Staff.objects.filter(is_active=True,is_own_client_suspect=True,created_at__year=year,created_at__month=month).order_by("created_at")
	else:
		client_list = Client_Staff.objects.filter(is_active=True,is_own_client_suspect=True).order_by("created_at")

	
	context = {
		'client_list': client_list,
		'menu':'client_own_suspect'
	}
	return render(request,template,context=context)

@login_required
def client_own_suspect_detail(request,uid_client_staff):
	template = 'admin/client/suspect/client_own_detail.html'
	
	client_staff_sus = Client_Staff.objects.filter(is_active=True,uid=uid_client_staff).first()
	

	if request.POST:
		action = request.POST['action']
		client = client_staff_sus.client
		if action == 'accept':
			print("accept")
			client.source = '0'
			client.source_detail_1 = 2
			client.source_detail_2 = None
			client_staff_sus.is_own_client_suspect = False
		else:
			print("reject")
			client_staff_sus.is_own_client_suspect = False
			
		client_staff_sus.updated_by = request.user
		client_staff_sus.updated_at = timezone.now()
		client_staff_sus.save()	
		client.save()

		return redirect(reverse('client-own-suspect-list'))
	
	context = {
		'client_staff_sus': client_staff_sus,
		'menu':'client_own_suspect',
	}
	return render(request,template,context=context)



class SuspectListView(ClientSuspectServerSideDatatableView):
	print('masuk class view')
	queryset = Client_Duplicate_Suspect.objects.filter(is_active=True,is_checked=False)

	columns = ['nama', 'magnet_created_at', 'phone_no', 'email', 'source', 'staff' , 'aecode', 'source_detail_2', 'medium', 'campaign', 'action']
	is_delete_data = True

@login_required
def client_suspect_list(request):
	now = timezone.now()
	month = request.GET.get('month')
	year = request.GET.get('year') or now.year


	staff = Staff.objects.filter(profile__user=request.user).first()
	template = 'admin/client/suspect/client_list.html'
	if month != "" and month != None:
		client_list = Client_Duplicate_Suspect.objects.filter(is_active=True,is_checked=False,created_at__year=year,created_at__month=month).order_by("-client_new__magnet_created_at").prefetch_related('client_old', 'client_new')
	else:
		client_list = Client_Duplicate_Suspect.objects.filter(is_active=True,is_checked=False).order_by("-client_new__magnet_created_at").prefetch_related('client_old', 'client_new')

	client_id_list = []
	for client in client_list:
		client_id_list.append(client.client_new.id)
	journey_register = Client_Journey.objects.filter(is_active=True,journal_type='registered',client__in=client_id_list)
	register_dict = {}
	for x in journey_register:
		register_dict[x.client] = x.created_at

	context = {
		'register_dict': register_dict,
		'client_list': client_list,
		'menu':'client_suspect'
	}
	return render(request,template,context=context)

@login_required
def client_suspect_history_list(request):
		
	template = 'admin/client/suspect/history_list.html'
	history_list = Client_Duplicate_Suspect_History.objects.filter(is_active=True).order_by("-updated_at")
	context = {
		'history_list': history_list,
		'menu':'client_suspect'
	}
	return render(request,template,context=context)

@login_required
@csrf_exempt
def client_suscpet_action(request):
	success = False
	if request.user.is_superuser == True:
		print('super user')
		try:
			with transaction.atomic():
				if request.is_ajax():
					
					response_data = request.body.decode()
					
					data_dict = json.loads(response_data)
					print(data_dict)

					if 'suspect_client_ids' in data_dict:
						print(data_dict['suspect_client_ids'])
						for str_suspect_client_id in data_dict['suspect_client_ids']:
							client_sus = Client_Duplicate_Suspect.objects.filter(id=int(str_suspect_client_id)).first()
							old_staff = Client_Staff.objects.filter(client=client_sus.client_old, is_active=True).first()
							new_staff = Client_Staff.objects.filter(client=client_sus.client_new, is_active=True).first()
							print('-----')	
							if old_staff != None:
								temp = old_staff
								old_staff = temp.staff
							if new_staff != None:
								temp = new_staff
								new_staff = temp.staff

							action = data_dict['action']
							if 'action_extra' in data_dict:
								action_extra = data_dict['action_extra']
							else:
								action_extra = ''
							client = client_sus.client_new
							client.is_suspect_bypass = True
							if action == 'accept':
								print("accept")
								client.is_suspect = False
								
								

								new_his = Client_Duplicate_Suspect_History()
								new_his.duplicate_suspect = client_sus
								new_his.action = "accepted"
								new_his.created_by = request.user

								new_his.save()
								
								old_client = client_sus.client_old
								old_client.is_suspect_bypass = True
								old_client.is_suspect = False
								old_client.updated_by = request.user
								old_client.save()

							elif action == 'take_right':
								client.is_suspect = False
								

								new_his = Client_Duplicate_Suspect_History()
								new_his.duplicate_suspect = client_sus
								new_his.action = "take_right"
								new_his.created_by = request.user

								new_his.save()

								old_client = client_sus.client_old
								old_client.is_active = False
								old_client.is_suspect = False
								old_client.is_suspect_bypass = True

								old_client.updated_by = request.user
								old_client.save()



							else:
								print("reject")
								new_his = Client_Duplicate_Suspect_History()
								new_his.duplicate_suspect = client_sus
								new_his.action = "rejected"
								new_his.created_by = request.user
								new_his.save()
								client.is_active = False
								if action_extra != None and action_extra != '' and action_extra == 'pribadi':
									client.source_detail_1 = 2

								rejected_client_followup = Client_Followup.objects.filter(is_active=True,client=client)
								for x in rejected_client_followup:
									follow = Client_Followup()
									follow.client = x.client
									follow.followup = x.followup
									follow.staff = x.staff
									follow.answer = x.answer
									follow.created_by = request.user
									follow.created_at = timezone.now() 
									follow.save()

								rejected_client_schedule = Client_Schedule.objects.filter(is_active=True,client=client)
								for x in rejected_client_schedule:
									schedule = Client_Schedule()
									schedule.schedule_date = x.schedule_date
									schedule.client = x.client
									schedule.staff = x.staff
									schedule.schedule_type = x.schedule_type
									schedule.status = x.status
									schedule.notes = x.notes
									schedule.answer = x.answer
									schedule.created_by = request.user
									schedule.created_at = timezone.now() 
									schedule.save()
								
								
								rejected_client_journey = Client_Journey.objects.filter(is_active=True,client=client)
								for x in rejected_client_journey:
									if x.journal_type != 'registered':
										journey = Client_Journey()
										journey.schedule_date = x.schedule_date
										journey.client = x.client
										journey.staff = x.staff
										journey.journal_type = x.journal_type
										journey.created_by = request.user
										journey.created_at = timezone.now() 
										journey.save()

							client.updated_by = request.user
							client.save()
							print('client saved !', client.is_suspect, client.nama)
							client_sus.is_checked = True

							client_sus.updated_by = request.user
							client_sus.updated_at = timezone.now()
							client_sus.save()
							print('client_sus saved!')

						success = True


		except IntegrityError as e:
			pritn(e)

	json_response = {
		'status' : success,
		'url': request.META.get('HTTP_REFERER')
	}
	return JsonResponse(json_response)

@login_required
def client_suspect_detail(request,id_client_sus):
	template = 'admin/client/suspect/client_detail.html'
	# cur_staff = Staff.objects.filter(profile__user=request.user).first()
	
	# client = Client.objects.filter(id=id_client).first()
	client_sus = Client_Duplicate_Suspect.objects.filter(id=id_client_sus).first()
	old_staff = Client_Staff.objects.filter(client=client_sus.client_old, is_active=True).first()
	new_staff = Client_Staff.objects.filter(client=client_sus.client_new, is_active=True).first()
	
	if old_staff != None:
		temp = old_staff
		old_staff = temp.staff
	if new_staff != None:
		temp = new_staff
		new_staff = temp.staff

	if request.POST:
		print(request.POST)
		action = request.POST['action']
		action_extra = request.POST['action_extra']
		client = client_sus.client_new
		client.is_suspect_bypass = True
		if action == 'accept':
			print("accept")
			client.is_suspect = False
			
			

			new_his = Client_Duplicate_Suspect_History()
			new_his.duplicate_suspect = client_sus
			new_his.action = "accepted"
			new_his.created_by = request.user

			new_his.save()

		elif action == 'take_right':
			client.is_suspect = False
			

			new_his = Client_Duplicate_Suspect_History()
			new_his.duplicate_suspect = client_sus
			new_his.action = "take_right"
			new_his.created_by = request.user

			new_his.save()

			old_client = client_sus.client_old
			old_client.is_active = False
			old_client.is_suspect = False
			old_client.is_suspect_bypass = True

			old_client.updated_by = request.user
			old_client.save()



		else:
			print("reject")
			new_his = Client_Duplicate_Suspect_History()
			new_his.duplicate_suspect = client_sus
			new_his.action = "rejected"
			new_his.created_by = request.user
			new_his.save()
			client.is_active = False
			if action_extra != None and action_extra != '' and action_extra == 'pribadi':
				client.source_detail_1 = 2

			rejected_client_followup = Client_Followup.objects.filter(is_active=True,client=client)
			for x in rejected_client_followup:
				follow = Client_Followup()
				follow.client = x.client
				follow.followup = x.followup
				follow.staff = x.staff
				follow.answer = x.answer
				follow.created_by = request.user
				follow.created_at = timezone.now() 
				follow.save()

			rejected_client_schedule = Client_Schedule.objects.filter(is_active=True,client=client)
			for x in rejected_client_schedule:
				schedule = Client_Schedule()
				schedule.schedule_date = x.schedule_date
				schedule.client = x.client
				schedule.staff = x.staff
				schedule.schedule_type = x.schedule_type
				schedule.status = x.status
				schedule.notes = x.notes
				schedule.answer = x.answer
				schedule.created_by = request.user
				schedule.created_at = timezone.now() 
				schedule.save()
			
			
			rejected_client_journey = Client_Journey.objects.filter(is_active=True,client=client)
			for x in rejected_client_journey:
				if x.journal_type != 'registered':
					journey = Client_Journey()
					journey.schedule_date = x.schedule_date
					journey.client = x.client
					journey.staff = x.staff
					journey.journal_type = x.journal_type
					journey.created_by = request.user
					journey.created_at = timezone.now() 
					journey.save()
			






		client.updated_by = request.user
		client.save()
		client_sus.is_checked = True

		client_sus.updated_by = request.user
		client_sus.updated_at = timezone.now()
		client_sus.save()
		print('client_sus id', client_sus.id, client_sus.is_checked)
		return redirect(reverse('client-suspect-list'))
	# history_followup = Client_Followup.objects.filter(client=client)
	# history_schedule = Client_Schedule.objects.filter(client=client)
	# history_journey = Client_Journey.objects.filter(client=client)

	
	context = {
		'client_sus': client_sus,
		'menu':'client_suspect',
		'old_staff': old_staff,
		'new_staff': new_staff,
		# 'client_exist':client_exist,
		# 'history_followup': history_followup,
		# 'history_schedule': history_schedule,
		# 'history_journey':history_journey,
		# 'history_schedule': client,
		# 'id_client':id_client,
	}
	return render(request,template,context=context)

@login_required	
def client_sync(request):
	sync_data_magnet()
	# check_user_deposit()
	 
	return redirect(reverse('client-list'))

class ClientListView(ServerSideDatatableView):
	queryset = Client.objects.filter(is_active=True, is_suspect=False, is_deposit=False).order_by('-created_at')

	columns = ['checkbox', 'nama', 'created_at', 'magnet_created_at', 'email', 'phone_no', 'source_detail_2', 'medium', 'campaign', 'id']
	is_delete_data = True

class DepositClientListView(ServerSideDatatableView):
	
	queryset = Client.objects.filter(is_active=True, is_suspect=False, is_deposit=True).order_by('-created_at')
	columns = ['nama', 'created_at', 'magnet_created_at', 'email', 'phone_no', 'source_detail_2', 'medium', 'campaign', 'id']
	is_delete_data = False


@login_required
def admin_client_list_ajax(request):
	followup_page = request.GET.get('page') or 1
	followup_client_list = Client.objects.filter(is_active=True, is_suspect=False, is_deposit=False).order_by("-created_at")
	followup_client_list_count = followup_client_list.count()

	followup_paginator = Paginator(followup_client_list, 100) # Show 25 contacts per page.
	followup_page_number = int(followup_page)
	followup_client_obj = followup_paginator.get_page(followup_client_list)
	response_data = {}
	response_data['recordsTotal'] = followup_client_list_count
	response_data['data'] = []
	counter = 1
	for followup_client in followup_client_obj:
		temp = {}
		temp['id'] = followup_client.id
		temp['no'] = counter
		temp['nama'] = followup_client.nama
		temp['nomor_telepon'] = followup_client.phone_no
		temp['email'] = followup_client.email
		temp['registered_date'] = str(followup_client.created_at)
		temp['source'] = followup_client.source
		temp['medium'] = followup_client.medium
		temp['campaign'] = followup_client.campaign
		temp['action'] = 'aksi'
		response_data['data'].append(temp)
	
	return JsonResponse(response_data)
@login_required
def admin_client_list(request):
	template = 'admin/client/admin_client_list.html'
	client_ajax_form = ClientAjaxForm(request.POST or None)
	client_color = {}
	client_color_text = {}

	followup_client_list = Client.objects.filter(is_active=True, is_suspect=False, is_deposit=False).order_by("-created_at")
	followup_client_list_count = followup_client_list.count()

	followup_paginator = Paginator(followup_client_list, 100) # Show 25 contacts per page.
	followup_page = request.GET.get('page') or 1
	followup_max_page = followup_paginator.num_pages
	followup_page_number = int(followup_page)
	followup_client_obj = followup_paginator.get_page(followup_client_list)


	if request.POST:
		print('masuk ajax')
		if client_ajax_form.is_valid():
			print('form ajax valid')
			query = client_ajax_form.cleaned_data['query']
			followup_page = client_ajax_form.cleaned_data['page'] or 1

			if query == '' and query == None:
				followup_client_list = Client.objects.filter(is_active=True, is_suspect=False, is_deposit=False).order_by("-created_at")
				followup_client_list_count = followup_client_list.count()
			else:
				followup_client_list = Client.objects.filter(is_active=True, is_suspect=False, is_deposit=False).order_by("-created_at")
				followup_paginator = Paginator(followup_client_list, 100) # Show 25 contacts per page.
				followup_page_number = int(followup_page)
				followup_client_obj = followup_paginator.get_page(followup_client_list)
				response_data = {}
				response_data['recordsTotal'] = followup_client_list_count
				response_data['data'] = []
				counter = 1
				for followup_client in followup_client_obj:
					temp = {}
					temp['id'] = followup_client.id
					temp['no'] = counter
					temp['nama'] = followup_client.nama
					temp['nomor_telepon'] = followup_client.phone_no
					temp['email'] = followup_client.email
					temp['registered_date'] = str(followup_client.created_at)
					temp['source'] = followup_client.source
					temp['medium'] = followup_client.medium
					temp['campaign'] = followup_client.campaign
					temp['action'] = 'aksi'
					response_data['data'].append(temp)

				return JsonResponse(response_data)
	context = {
		'client_ajax_form': client_ajax_form,
		'followup_client_obj': followup_client_obj,
		'followup_max_page': followup_max_page,

		'menu':'client',
		# 'staff_level':staff.staff_level.level,
	}

	return render(request,template,context=context)


@login_required
def client_list(request):
		
	staff = Staff.objects.filter(profile__user=request.user).first()

	client_color = {}
	client_color_text = {}
	color_dict = {
		'008000': 'Hijau',
		'ff0000': 'Merah',
		'e7ff00': 'Kuning',
		'000000': 'Hitam',
		'0066ff': 'Biru'
	}

	if not request.user.is_superuser and not staff.staff_level.level < 2:
		template = 'admin/client/client_list.html'
		# staff = Staff.objects.filter(is_active=True, profile__user__id=request.user.id).first()

		client_staff_list = Client_Staff.objects.filter(staff__id=staff.id, is_active=True, client__is_suspect=False).order_by('-created_at','-client__magnet_created_at').prefetch_related('client')
		client_list = Client.objects.none()

		
		
		for client_staff in client_staff_list:
			if client_staff.client.id not in client_color:
				client_color[client_staff.client.id] = client_staff.color
				client_color_text[client_staff.client.id] = color_dict[client_staff.color]
			

	else:
		template = 'admin/client/admin_client_list.html'
		client_list = Client.objects.filter(is_active=True, is_suspect=False).order_by("-created_at")
		client_staff_list = Client_Staff.objects.none()

	form_color = ColorForm(None)
	
	# messages.success(request, 'Profile details updated.')
	context = {
		'client_staff_list': client_staff_list,
		'form_color' : form_color,
		'menu':'client',
		'client_color':client_color,
		'client_color_text' : client_color_text,
		# 'staff_level':staff.staff_level.level,
	}
	return render(request,template,context=context)

@login_required
def client_add(request):
	template = 'admin/client/client_add.html'
	client_form = ClientForm(request.POST or None)

	if request.POST:
		try:
			with transaction.atomic():

				if client_form.is_valid() :
					
					client = client_form.save(commit=False)
					client.created_by = request.user
					client.save()

					staff = Staff.objects.filter(is_active=True, profile__user=request.user).first()
					if staff != None:
						client_staff = Client_Staff()
						client_staff.client = client
						client_staff.staff = staff
						client_staff.created_by = client_staff.updated_by = request.user
						client_staff.save()


					return redirect(reverse('client-list'))
				else:
					print(client_form.errors)

		except IntegrityError as e:
			print(e)

	context = {
		'client_form': client_form,
		'menu':'client',
	}
	return render(request,template,context=context)

@login_required
def client_edit(request,id_client):
	template = 'admin/client/client_add.html'

	client = Client.objects.filter(id=id_client,is_active=True).first()
	
	client_form = ClientForm(request.POST or None,instance=client, )


	if request.POST:
		try:
			with transaction.atomic():

				if client_form.is_valid() :
					
					client = client_form.save(commit=False)
					client.updated_by = request.user
					client.updated_at = timezone.now()

					client_aecode = client_form.cleaned_data['aecode']

					#get existing client_staff
					client_staff = Client_Staff.objects.filter(client=client, is_active=True).first()
					if client_staff != None:
						old_staff = client_staff.staff
						if old_staff != None:
							if old_staff.aecode != client_aecode:
								selected_staff = Staff.objects.filter(aecode=client_aecode, is_active=True).first()
								new_client_staff = Client_Staff()
								new_client_staff.client = client
								new_client_staff.staff = selected_staff
								new_client_staff.updated_by = new_client_staff.created_by = request.user
								new_client_staff.save()

								old_client_journey_list = Client_Journey.objects.filter(client=client, staff=old_staff, is_active=True)
								for old_client_journey in old_client_journey_list:
									new_client_journey = Client_Journey()
									new_client_journey.client = client
									new_client_journey.staff = selected_staff
									new_client_journey.journal_type = old_client_journey.journal_type
								new_client_journey.updated_by = new_client_journey.created_by = request.user
								new_client_journey.save()
					else:
						selected_staff = Staff.objects.filter(aecode=client_aecode, is_active=True).first()
						new_client_staff = Client_Staff()
						new_client_staff.client = client
						new_client_staff.staff = selected_staff
						new_client_staff.updated_by = new_client_staff.created_by = request.user
						new_client_staff.save()
					


					client.aecode = client_aecode
					client.save()

					return redirect(reverse('client-list'))
				else:
					print(client_form.errors)

		except IntegrityError as e:
			print(e)

	context = {
		'client_form': client_form,
		'menu':'client',
	}
	return render(request,template,context=context)

@login_required
def client_edit_color(request,id_client, color_str):
	template = 'admin/client/client_add.html'

	try:
		with transaction.atomic():
			# client = Client.objects.filter(id=id_client,is_active=True).first()
			staff = Staff.objects.filter(profile__user__id=request.user.id).first()
			# print('id_client',id_client,' staff.id',staff.id)
			client_staff = Client_Staff.objects.filter(client__id=id_client, is_active=True).first()
			client_staff.color = color_str
			print(color_str,'color_str')
			client_staff.updated_by = request.user

			client_staff.save()
			# print(client_staff.color, color_str)
	
	except IntegrityError as e:
		print(e)

	return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def client_set_hot(request,id_client, action):


	try:
		with transaction.atomic():
			
			staff = Staff.objects.filter(profile__user__id=request.user.id).first()
			client_staff = Client_Staff.objects.filter(client__id=id_client, is_active=True).first()

			if action == "true":
				client_staff.is_hot = True
			else:
				client_staff.is_hot = False
			client_staff.updated_by = request.user
			client_staff.save()
	
	except IntegrityError as e:
		print(e)

	return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def request_own(request,id_client):
	template = 'admin/client/client_detail_list.html'
	cur_staff = Staff.objects.filter(profile__user=request.user).first()
	client = Client.objects.filter(id=id_client).first()
	client_staff = Client_Staff.objects.filter(is_active=True,client=client).first()

	client_staff.is_own_client_suspect = True
	client_staff.updated_at = timezone.now()
	client_staff.updated_by = request.user
	client_staff.save()

	messages.success(request, "Request Berhasil")
	return redirect(reverse('client-detail-list', kwargs={'id_client': id_client,}))

@login_required
def detail_list(request,id_client):
	template = 'admin/client/client_detail_list.html'
	cur_staff = Staff.objects.filter(profile__user=request.user).first()
	client = Client.objects.filter(id=id_client).first()
	client_staff = Client_Staff.objects.filter(is_active=True,client=client,staff=cur_staff,is_own_client_suspect=False).exclude(client__source='0').first()
	
	

	history_followup = Client_Followup.objects.filter(client=client)
	history_schedule = Client_Schedule.objects.filter(client=client)
	history_journey = Client_Journey.objects.filter(client=client).order_by('-created_at')
	dict_total = {}
	if client.magnet_id == None  or client.magnet_id == '':
		print('tidak ada magnet id')
		selected_client_magnet_id = get_client_magnet_id(client)
		if selected_client_magnet_id != None:
			client.magnet_id = selected_client_magnet_id
			client.save()

	if client.magnet_id != None:
		client_position_list,pos_detail = get_client_position(client.id)
	

	
		client_eq_bal = get_login_trades(client.id)
		if client_position_list != None and client_position_list != []:
			for x in client_position_list:
				dict_total[x[1]] = x[16] + x[17]

			client_eq_bal = client_eq_bal['data'][0] if 'data' in client_eq_bal else []
		else:
			client_eq_bal = []
	
	context = {
		'client': client,
		'client_staff':client_staff,
		'history_followup': history_followup,
		'history_schedule': history_schedule,
		'history_journey':history_journey,
		'client_position_list':client_position_list,
		'pos_detail':pos_detail,
		'client_eq_bal':client_eq_bal,
		# 'history_schedule': client,
		'dict_total':dict_total,
		'id_client':id_client,
	}
	return render(request,template,context=context)

@login_required
@csrf_exempt
def client_delete(request):
	print('client_delete')
	success = False
	if request.user.is_superuser == True:
		print('super user')
		try:
			with transaction.atomic():
				if request.is_ajax():
					print('halo request post')
					print (request.body)
					print(request.body.decode())
					response_data = request.body.decode()
					print(type(response_data))
					data_dict = json.loads(response_data)
					print(data_dict)
					if 'client_ids' in data_dict:
						print(data_dict['client_ids'])
						for str_client_id in data_dict['client_ids']:
							id_client = int(str_client_id)
					

							client = Client.objects.filter(id=id_client).first()
							client.is_active = False
							client.updated_by = request.user
							client.updated_at = timezone.now()
							client.save()

							client_staff_list = Client_Staff.objects.filter(client=client, is_active = True)
							for client_staff in client_staff_list:
								client_staff.is_active = False
								client_staff.updated_by = request.user
								client_staff.updated_at = timezone.now()
								client_staff.save()

							client_journey_list = Client_Journey.objects.filter(client=client, is_active=True)
							for client_journey in client_journey_list:
								client_journey.is_active = False
								client_journey.updated_by = request.user
								client_journey.updated_at = timezone.now()
								client_journey.save()

						success = True


		except IntegrityError as e:
			pritn(e)

	json_response = {
		'status' : success,
		'url': request.META.get('HTTP_REFERER')
	}
	return JsonResponse(json_response)

@login_required
def client_followup_list(request,id_client):
	template = 'admin/client/client_followup_list.html'
	cur_staff = Staff.objects.filter(profile__user=request.user).first()
	client_followup_list = Client_Followup.objects.filter(is_active=True,staff=cur_staff,client__id=id_client)

	
	context = {
		'client_followup_list': client_followup_list,
		'id_client':id_client,
	}
	return render(request,template,context=context)


@login_required
def cleanhtml(raw_html):
	  cleantext = re.sub(CLEANR, '', raw_html)
	  return cleantext


@login_required
def client_schedule_list(request, client_id):
	template = 'admin/client/client_schedule/list.html'
	staff = Staff.objects.filter(profile__user=request.user).first()
	client = Client.objects.filter(id=client_id).first()
	client_schedule_list = Client_Schedule.objects.filter(client__id=client.id, is_active=True).order_by('schedule_date')
	

	
	context = {
		'client_schedule_list': client_schedule_list,
		'client':client,
	}
	return render(request,template,context=context)

@login_required
def client_schedule_add(request, client_id):
	template = 'admin/client/client_schedule/add.html'
	staff = Staff.objects.filter(profile__user=request.user).first()
	client = Client.objects.filter(id=client_id).first()
	datetime_form = DateTimeForm(request.POST or None,initial={' schedule_date': None})
	form = ClientScheduleForm(request.POST, prefix='normal-form')
	
	if request.POST:
		try:
			with transaction.atomic():
				if form.is_valid() and datetime_form.is_valid():
					client_schedule = form.save(commit=False)
					schedule_date = datetime.strptime(datetime_form.cleaned_data['schedule_date'], '%Y-%m-%d %H:%M')
					client_schedule.schedule_date = schedule_date
					client_schedule.staff = staff
					client_schedule.client = client
					client_schedule.created_by = client_schedule.updated_by = request.user
					client_schedule.save()
					
					message_str = ('Schedule for %s has been registered'%(client.nama, ) )
					messages.success(request, message_str)
					message_str = ('Staff %s (%s) has been unlocked'%(staff.profile.full_name, staff.staff_level.level_name) )
			
					# start_process.apply_async(_id='eta-testing')
					# schedule_date_reminder = timezone.now()
					# reminder_date = schedule_date - timedelta(hours=0, minutes=1)
					# schedule_date_reminder = schedule_date_reminder.replace(day=reminder_date.day,month=reminder_date.month,year=reminder_date.year,hour=reminder_date.hour,minute=reminder_date.minute,second=reminder_date.second) 
					# print(schedule_date_reminder,"schedule_date_reminder",timezone.now(),"timezone.now()",schedule_date_reminder-timezone.now())
					print(make_aware(schedule_date - timedelta(hours=0, minutes=5))-timezone.now())
					# make_aware(schedule_date - timedelta(days=30))
					# make_aware(schedule_date - timedelta(days=14))
					# make_aware(schedule_date - timedelta(days=7))
					# make_aware(schedule_date - timedelta(days=1))
					# start_process.apply_async(_id='eta-testing',eta=make_aware(schedule_date))

					return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
					
				else:
					print("else")
					cleantext = re.sub(CLEANR, '', str(form.errors))
					print(cleantext,str(datetime_form.errors))
					messages.error(request, cleantext)
		except IntegrityError as e:
			print("masuk integrity")
			messages.error(request, e)

	context = {}
	context['datetime_form'] = datetime_form
	context['form'] = form
	context['staff'] = staff
	context['client'] = client

	return render(request,template,context=context)

@login_required
def client_schedule_update(request, client_schedule_uid):
	
	template = 'admin/client/client_schedule/update.html'
	client_schedule = Client_Schedule.objects.filter(uid=client_schedule_uid).prefetch_related('client').first()
	print(client_schedule)
	client = client_schedule.client

	staff = Staff.objects.filter(profile__user=request.user).first()
	datetime_form = DateTimeForm(request.POST or None,initial={'schedule_date': client_schedule.schedule_date.strftime('%Y-%m-%d %H:%M')})
	# datetime_form = DateTimeForm(request.POST, schedule_date=client_schedule.schedule_date.strftime('%Y-%m-%d %H:%M'))
	form = ClientScheduleForm(request.POST or None, instance=client_schedule, prefix='client-schedule-form')

	if request.POST:
		try:
			with transaction.atomic():
				if form.is_valid() and datetime_form.is_valid():
					print('valid')
					client_schedule = form.save(commit=False)
					schedule_date = datetime.strptime(datetime_form.cleaned_data['schedule_date'], '%Y-%m-%d %H:%M')
					client_schedule.schedule_date = schedule_date
					client_schedule.updated_by = request.user
					client_schedule.save()
					
					message_str = ('Schedule for %s has been registered'%(client.nama, ) )
					messages.success(request, message_str)
					message_str = ('Staff %s (%s) has been unlocked'%(staff.profile.full_name, staff.staff_level.level_name) )
			

			
					return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
					
				else:
					print('eroors', form.errors, datetime_form.errors)
					cleantext = re.sub(CLEANR, '', str(form.errors))
					print(form.errors, datetime_form.errors)
					messages.error(request, cleantext)
		except IntegrityError as e:
			messages.error(request, e)

	context = {}
	context['datetime_form'] = datetime_form
	context['form'] = form
	context['staff'] = staff
	context['client'] = client

	return render(request,template,context=context)

@login_required
def client_journey_add(request, client_id, journey_type):
	data = {}
	data['status'] = False
	try:
		with transaction.atomic():
			staff = Staff.objects.filter(profile__user__id=request.user.id).first()
			print(staff.id, '<----')
			client = Client.objects.filter(id=client_id).first()
			if create_client_journal(request, staff, client, journey_type) == True:
				data['status'] = True

	except IntegrityError as e:
		print(e)
	return JsonResponse(data)

@login_required
def feedback_list(request):
	template = 'admin/client/client_feedback/list.html'

	total_top = Settings.objects.filter(is_active=True).first().top_client_feedback or 3
	staff = Staff.objects.filter(profile__user__id=request.user.id, is_active=True).first()
	# sub_staff_list = Staff.objects.filter(staff_parent__id=staff.id, is_active=True)
	# if sub_staff_list.count() == 0 >:

	if request.user.is_superuser or staff.staff_level.level < 2 :
		client_feedback_list = Client_Followup.objects.filter(is_active=True).prefetch_related('followup', 'client', 'staff').order_by('-created_at')[:total_top]
	else:
		child_staff = Staff.objects.filter(staff_parent = staff,is_active=True)
		if child_staff is None:
			client_feedback_list = Client_Followup.objects.filter(is_active=True, staff=staff).prefetch_related('followup', 'client', 'staff').order_by('-created_at')[:total_top]
		else:
			client_feedback_list = Client_Followup.objects.filter(is_active=True, staff__in=child_staff.values_list('id',flat=True)).prefetch_related('followup', 'client', 'staff').order_by('-created_at')[:total_top]
			

	dict_client_res = {}
	for client_res in client_feedback_list:
		if client_res.followup not in dict_client_res:
			dict_client_res[client_res.followup] = 0
		dict_client_res[client_res.followup] += 1

	print(dict_client_res)

	dict_client_res_sorted = {}
	# dict_client_res = sorted(dict_client_res.items(), key=lambda x: x[1], reverse=True)
	top_response = {}
	counter = 0
	for i in sorted(dict_client_res.items(), key=lambda x: x[1], reverse=True):
		dict_client_res_sorted[i[0]] = i[1]
		if counter < total_top:
			top_response[i[0]] = i[1]
			counter+=1


	

	# [(dict_client_res_sorted[key] = value) for (key, value) in sorted(dict_client_res.items(), key=lambda x: x[1])]
	# print(dict_client_res)



	context = {}
	context['client_feedback_list'] = client_feedback_list
	context['dict_client_res'] = dict_client_res_sorted
	context['top_response'] = top_response
	context['staff'] = staff
	context['total_top'] = total_top

	context['menu'] = 'client_feedback'


	return render(request,template,context=context)

@login_required
def client_position(request, client_id):
	client = Client.objects.filter(id=client_id).first()
	print(client.magnet_id)
	client_position_list = get_client_position(client.id)
	print(client)

	context = {
		'client_position_list': client_position_list,
		'client': client,
	}
	template = 'admin/client/position.html'
	# template = 'admin/client/client_schedule/list.html'
	return render(request, template, context)

@login_required
def client_position_history(request, client_id):
	client = Client.objects.filter(id=client_id).first()
	print(client.magnet_id)
	from_date = request.GET.get('from') or None
	to_date = request.GET.get('to') or None

	client_position_history = get_login_trades_history(client.id,from_date,to_date)
	print("client_position_history",client_position_history)

	total_in = 0
	total_out = 0
	for x in client_position_history['data']:
		if x['entry'] == "open":
			total_in += Decimal(x['lot'])
		else:
			total_out += Decimal(x['lot'])


	context = {
		'client_position_history': client_position_history['data'] if 'data' in client_position_history else [] ,
		'client': client,
		'from_date':from_date,
		'to_date':to_date,
		'total_in':total_in,
		'total_out':total_out,


	}
	template = 'admin/client/position_history.html'
	# template = 'admin/client/client_schedule/list.html'
	return render(request, template, context)
	
	

